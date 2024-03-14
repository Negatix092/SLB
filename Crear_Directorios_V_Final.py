# Version 2024-02-29 
import pandas as pd
import pyodbc
import os
import re
import shutil
import datetime
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Año de interés definido aquí
now = datetime.datetime.now()
año_interes = now.year
mes_actual = now.month

# Ajustar el trimestre de interés basado en el mes actual
if mes_actual in [4, 7, 10, 1]:  # Meses siguientes al cierre de cada trimestre
    trimestre_actual = (mes_actual - 1) // 3
    if mes_actual == 1:  # Enero, ajustar al último trimestre del año anterior
        trimestre_actual = 4
        año_interes = now.year - 1
else:  # Para cualquier otro mes, procesar el trimestre actual
    trimestre_actual = (mes_actual - 1) // 3 + 1

estado_documentos = {} #diccionario para almacenar el estado de los documentos

# Información de conexión

server = 'ec0038app05'
database = 'SHAYA'
username = 'python_user'
password = 'python_user'

conn_str = f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'

directorio_base_copia = rf'\\dir.slb.com\NSA\SAM_Collaborate\EC0037\00_Well_File'

# Función para obtener los DataFrames
def obtener_dataframe(query, conn):
    return pd.read_sql(query, conn)

# Función para filtrar el DataFrame por trimestre y WO_NUMBER (para el caso de Capex y Opex)
def filtrar_por_trimestreCapex(df, meses):
    return df[df['END_WO'].dt.month.isin(meses) & df['WO_NUMBER'].notna() & (df['WO_NUMBER'] != 0)].copy()

# Función para filtrar el DataFrame por trimestre para el caso de CPI
def filtrar_por_trimestreCPI_Opex(df, meses):
    return df[df['END_WO'].dt.month.isin(meses)].copy()

# Función para obtener rutas de origen desde el directorio_base_copia usando el nombre de los pozos
def obtener_ruta_origen(pozo):
    # Defino un diccionario con los prefijos de los pozos y sus carpetas correspondientes
    carpetas = {'ANC': 'ANACONDA', 'ANR': 'ANURA', 'ACA': 'AUCA CENTRAL', 'ACS': 'AUCA SUR', 'BOA': 'BOA',
                'CHE': 'CHONTA ESTE', 'CHS': 'CHONTA SUR', 'CG': 'CONGA', 'CNO': 'CONONACO', 'CLB': 'CULEBRA',
                'PTL': 'PITALALA', 'RMY': 'RUMIYACU', 'TTS': 'TORTUGA', 'YCA': 'YUCA', 'YLB': 'YULEBRA'}
    
    carpetas_especiales = {'CGSA': os.path.join('CONGA', 'CONGA SUR')}

    # Primero verifica los casos especiales
    if pozo.startswith('CGSA'):
        return os.path.join(directorio_base_copia, carpetas_especiales['CGSA'], pozo)

    # Luego verifica los casos normales
    for prefijo, carpeta in carpetas.items():
        if pozo.startswith(prefijo):
            return os.path.join(directorio_base_copia, carpeta, pozo)
    
    return None

# Funcion para obtener un directorio utilizando el numero de workover (para el caso de Capex y Opex)
def obtener_directorio_por_workover(ruta_base, wo_number):
    
    wo_number = str(wo_number)
    # Crear la expresión regular
    regex = re.compile(rf'WO.*{wo_number}')

    # Buscar en la carpeta base
    for root, dirs, _ in os.walk(ruta_base):
        for dir in dirs:
            if regex.search(dir):
                return os.path.join(root, dir)
    
    return None

def obtener_directorio_paraCPI(ruta_base):
    for root, dirs, _ in os.walk(ruta_base):
        for dir in dirs:
            if dir.startswith('CPI'):
                return os.path.join(root, dir)
    return None

# Implementación de la función para encontrar una carpeta por prefijo (para carpetas numeradas en los directorios)
def encontrar_carpeta_por_prefijo(ruta_base, prefijo):
    carpetas = next(os.walk(ruta_base))[1]
    for carpeta in carpetas:
        if carpeta.startswith(prefijo):
            return os.path.join(ruta_base, carpeta)
    return None

# Función modificada para obtener el directorio específico AIS/ARS
def obtener_directorio_AIS(ruta_base, pozo, wo_number):
    wo_number = str(wo_number)
    # Si el pozo termina en 'I' o 'H', ignorar ese último caracter (debido a que en los AIS/ARS no siempre ponen el nombre como en la base de datos)
    if pozo.endswith('I') or pozo.endswith('H'):
        pozo = pozo[:-1]
    # Crear la expresión regular 
    regex = re.compile(rf'.*{pozo}.*WO.*{wo_number}.*')

    for root, dirs, _ in os.walk(ruta_base):
        for dir in dirs:
            if regex.search(dir):
                return os.path.join(root, dir)
    
    return None

def obtener_directorio_ARS(ruta_base, pozo):
    # Si el pozo termina en 'I' o 'H', ignorar ese último caracter (debido a que en los AIS/ARS no siempre ponen el nombre como en la base de datos)
    if pozo.endswith('I') or pozo.endswith('H'):
        pozo = pozo[:-1]
    # Crear la expresión regular    
    regex = re.compile(rf'.*{pozo}.*ABANDONO.*')

    for root, dirs, _ in os.walk(ruta_base):
        for dir in dirs:
            if regex.search(dir):
                return os.path.join(root, dir)
    
    return None

def obtener_directorio_AIS_CPI(ruta_base, pozo):
    # Si el pozo termina en 'I' o 'H', ignorar ese último caracter (debido a que en los AIS/ARS no siempre ponen el nombre como en la base de datos)
    if pozo.endswith('I') or pozo.endswith('H'):
        pozo = pozo[:-1]
    # Crear la expresión regular    
    regex = re.compile(rf'.*{pozo}.*')

    for root, dirs, _ in os.walk(ruta_base):
        for dir in dirs:
            if regex.search(dir):
                return os.path.join(root, dir)
    
    return None

# Funciones para copiar archivos
# estado_documentos es un diccionario que almacena el estado de los documentos copiados

# Caso en que se desea copiar un archivo pdf (Soporte para Inicio de Trabajos) para Capex y Opex
def copiar_pdf(pozo, wo_number, ruta_destino, tipo_pozo, abandono):

    ruta_base_origen = obtener_ruta_origen(pozo)
    ruta_origen_workover = obtener_directorio_por_workover(ruta_base_origen, wo_number)

    estado_documento = 'X' # Estado por defecto
    if 'adecua' in abandono.lower():
        pozo = f'{pozo} ADECUACION'
    if pozo not in estado_documentos:
        estado_documentos[pozo] = {}
    if tipo_pozo not in estado_documentos[pozo]:
        estado_documentos[pozo][tipo_pozo] = {}
    if wo_number not in estado_documentos[pozo][tipo_pozo]:
        estado_documentos[pozo][tipo_pozo][wo_number] = {}

    if ruta_origen_workover:
        # Encuentra la primera subcarpeta que comienza con "1."
        ruta_primera_subcarpeta = encontrar_carpeta_por_prefijo(ruta_origen_workover, '1.')
        if ruta_primera_subcarpeta:
            # Encuentra la segunda subcarpeta que comienza con "3." dentro de la primera subcarpeta
            ruta_segunda_subcarpeta = encontrar_carpeta_por_prefijo(ruta_primera_subcarpeta, '3.')
            if ruta_segunda_subcarpeta:
                # Encuentra la tercera subcarpeta que comienza con "1." dentro de la segunda subcarpeta
                ruta_tercera_subcarpeta = encontrar_carpeta_por_prefijo(ruta_segunda_subcarpeta, '1.')
                if ruta_tercera_subcarpeta:
                    for archivo in os.listdir(ruta_tercera_subcarpeta):
                        ruta_origen = os.path.join(ruta_tercera_subcarpeta, archivo)
                        if os.path.isfile(ruta_origen) and (archivo.endswith('.pdf') or archivo.endswith('.PDF') or "notificaci" in archivo.lower()):
                            shutil.copy(ruta_origen, ruta_destino)
                            estado_documento = 'OK'

    # Actualizar el estado del documento en el diccionario para el pozo correspondiente                
    estado_documentos[pozo][tipo_pozo][wo_number]['Notificación'] = estado_documento

# Caso en que se desea copiar un archivo pdf (Soporte para Inicio de Trabajos) para CPI
def copiar_pdf_cpi(pozo, ruta_destino):
    ruta_base_origen = obtener_ruta_origen(pozo)
    ruta_cpi = obtener_directorio_paraCPI(ruta_base_origen)

    estado_documento = 'X' # Estado por defecto
    if pozo not in estado_documentos:
        estado_documentos[pozo] = {'CPI': {}}
    
    if ruta_cpi:
        # Encuentra la primera subcarpeta que comienza con "1."
        ruta_primera_subcarpeta = encontrar_carpeta_por_prefijo(ruta_cpi, '1.')
        if ruta_primera_subcarpeta:
            # Encuentra la segunda subcarpeta que comienza con "3." dentro de la primera subcarpeta
            ruta_segunda_subcarpeta = encontrar_carpeta_por_prefijo(ruta_primera_subcarpeta, '3.')
            if ruta_segunda_subcarpeta:
                # Encuentra la tercera subcarpeta que comienza con "1." dentro de la segunda subcarpeta
                ruta_tercera_subcarpeta = encontrar_carpeta_por_prefijo(ruta_segunda_subcarpeta, '1.')
                if ruta_tercera_subcarpeta:
                    for archivo in os.listdir(ruta_tercera_subcarpeta):
                        ruta_origen = os.path.join(ruta_tercera_subcarpeta, archivo)
                        if os.path.isfile(ruta_origen) and (archivo.endswith('.pdf') or archivo.endswith('.PDF') or "notificaci" in archivo.lower()):
                            shutil.copy(ruta_origen, ruta_destino)
                            estado_documento = 'OK'

    # Actualizar el estado del documento en el diccionario para el pozo correspondiente                        
    estado_documentos[pozo]['CPI']['Notificación'] = estado_documento

# Caso en que se desea copiar un archivo word (Prognosis y Programas) para Capex y Opex
def copiar_prognosis_capex(pozo, wo_number, ruta_destino, tipo_pozo, abandono):
    ruta_base_origen = obtener_ruta_origen(pozo)
    ruta_origen_workover = obtener_directorio_por_workover(ruta_base_origen, wo_number)

    estado_documento = 'X' # Estado por defecto
    if 'adecua' in abandono.lower():
        pozo = f'{pozo} ADECUACION'
    if pozo not in estado_documentos:
        estado_documentos[pozo] = {}
    if tipo_pozo not in estado_documentos[pozo]:
        estado_documentos[pozo][tipo_pozo] = {}
    if wo_number not in estado_documentos[pozo][tipo_pozo]:
        estado_documentos[pozo][tipo_pozo][wo_number] = {}

    if ruta_origen_workover:
        # Encuentra la primera subcarpeta que comienza con "1."
        ruta_primera_subcarpeta = encontrar_carpeta_por_prefijo(ruta_origen_workover, '1.')
        if ruta_primera_subcarpeta:
            # Encuentra la segunda subcarpeta que comienza con "2." dentro de la primera subcarpeta
            ruta_segunda_subcarpeta = encontrar_carpeta_por_prefijo(ruta_primera_subcarpeta, '2.')
            if ruta_segunda_subcarpeta:
                for archivo in os.listdir(ruta_segunda_subcarpeta):
                    ruta_origen = os.path.join(ruta_segunda_subcarpeta, archivo)
                    if os.path.isfile(ruta_origen) and (archivo.endswith('.docx') or archivo.endswith('.doc') or archivo.endswith('.DOCX') or archivo.endswith('.DOC') or "propuesta" in archivo.lower()):
                        shutil.copy(ruta_origen, ruta_destino)
                        estado_documento = 'OK'

    # Actualizar el estado del documento en el diccionario para el pozo correspondiente       
    estado_documentos[pozo][tipo_pozo][wo_number]['Propuesta Técnica'] = estado_documento

# Caso en que se desea copiar un archivo word (Prognosis y Programas) para CPI
def copiar_prognosis_cpi(pozo, ruta_destino):
    ruta_base_origen = obtener_ruta_origen(pozo)
    ruta_cpi = obtener_directorio_paraCPI(ruta_base_origen)

    estado_documento = 'X' # Estado por defecto
    if pozo not in estado_documentos:
        estado_documentos[pozo] = {'CPI': {}}

    if ruta_cpi:
        # Encuentra la primera subcarpeta que comienza con "1."
        ruta_primera_subcarpeta = encontrar_carpeta_por_prefijo(ruta_cpi, '1.')
        if ruta_primera_subcarpeta:
            # Encuentra la segunda subcarpeta que comienza con "2." dentro de la primera subcarpeta
            ruta_segunda_subcarpeta = encontrar_carpeta_por_prefijo(ruta_primera_subcarpeta, '2.')
            if ruta_segunda_subcarpeta:
                for archivo in os.listdir(ruta_segunda_subcarpeta):
                    ruta_origen = os.path.join(ruta_segunda_subcarpeta, archivo)
                    if os.path.isfile(ruta_origen) and (archivo.endswith('.docx') or archivo.endswith('.doc') or archivo.endswith('.DOCX') or archivo.endswith('.DOC') or "propuesta" in archivo.lower()):
                        shutil.copy(ruta_origen, ruta_destino)
                        estado_documento = 'OK'

    # Actualizar el estado del documento en el diccionario para el pozo correspondiente       
    estado_documentos[pozo]['CPI']['Propuesta Técnica'] = estado_documento

# Caso en que se desea copiar todos los archivos pdf (Reportes Diarios de Ejecución) para Capex y Opex
def copiar_reportes_diarios(pozo, wo_number, ruta_destino, tipo_pozo, abandono):
    num_archivos_copiados = 0
    ruta_base_origen = obtener_ruta_origen(pozo)
    ruta_origen_workover = obtener_directorio_por_workover(ruta_base_origen, wo_number)

    estado_documento = 'X' # Estado por defecto
    if 'adecua' in abandono.lower():
        pozo = f'{pozo} ADECUACION'
    if pozo not in estado_documentos:
        estado_documentos[pozo] = {}
    if tipo_pozo not in estado_documentos[pozo]:
        estado_documentos[pozo][tipo_pozo] = {}
    if wo_number not in estado_documentos[pozo][tipo_pozo]:
        estado_documentos[pozo][tipo_pozo][wo_number] = {}

    if ruta_origen_workover:
        # Encuentra la primera subcarpeta que comienza con "2."
        ruta_primera_subcarpeta = encontrar_carpeta_por_prefijo(ruta_origen_workover, '2.')
        if ruta_primera_subcarpeta:
            # Encuentra la segunda subcarpeta que comienza con "3." dentro de la primera subcarpeta
            ruta_segunda_subcarpeta = encontrar_carpeta_por_prefijo(ruta_primera_subcarpeta, '3.')
            if ruta_segunda_subcarpeta:
                for archivo in os.listdir(ruta_segunda_subcarpeta):
                    if archivo.endswith('.pdf') or archivo.endswith('.PDF'):
                        shutil.copy(os.path.join(ruta_segunda_subcarpeta, archivo), ruta_destino)
                        num_archivos_copiados += 1
                        estado_documento = f'OK, {num_archivos_copiados} Reportes Diarios'

    # Actualizar el estado del documento en el diccionario para el pozo correspondiente            
    estado_documentos[pozo][tipo_pozo][wo_number]['Reportes Diarios'] = estado_documento

# Caso en que se desea copiar todos los archivos pdf (Reportes Diarios de Ejecución) para CPI
def copiar_reportes_diarios_cpi(pozo, ruta_destino):
    num_archivos_copiados = 0
    ruta_base_origen = obtener_ruta_origen(pozo)
    ruta_cpi = obtener_directorio_paraCPI(ruta_base_origen)

    estado_documento = 'X' # Estado por defecto
    if pozo not in estado_documentos:
        estado_documentos[pozo] = {'CPI': {}}

    if ruta_cpi:
        # Encuentra la primera subcarpeta que comienza con "2."
        ruta_primera_subcarpeta = encontrar_carpeta_por_prefijo(ruta_cpi, '2.')
        if ruta_primera_subcarpeta:
            # Encuentra la segunda subcarpeta que comienza con "3." dentro de la primera subcarpeta
            ruta_segunda_subcarpeta = encontrar_carpeta_por_prefijo(ruta_primera_subcarpeta, '3.')
            if ruta_segunda_subcarpeta:
                for archivo in os.listdir(ruta_segunda_subcarpeta):
                    if archivo.endswith('.pdf') or archivo.endswith('.PDF'):
                        shutil.copy(os.path.join(ruta_segunda_subcarpeta, archivo), ruta_destino)
                        num_archivos_copiados += 1
                        estado_documento = f'OK, {num_archivos_copiados} Reportes Diarios'

    # Actualizar el estado del documento en el diccionario para el pozo correspondiente         
    estado_documentos[pozo]['CPI']['Reportes Diarios'] = estado_documento

# Caso en que se desea copiar un archivo excel (Sumarios Finales) para Capex y Opex
def copiar_sumario(pozo, wo_number, ruta_destino, tipo_pozo, abandono):
    ruta_base_origen = obtener_ruta_origen(pozo)
    ruta_origen_workover = obtener_directorio_por_workover(ruta_base_origen, wo_number)

    estado_documento = 'X' # Estado por defecto
    if 'adecua' in abandono.lower():
        pozo = f'{pozo} ADECUACION'
    if pozo not in estado_documentos:
        estado_documentos[pozo] = {}
    if tipo_pozo not in estado_documentos[pozo]:
        estado_documentos[pozo][tipo_pozo] = {}
    if wo_number not in estado_documentos[pozo][tipo_pozo]:
        estado_documentos[pozo][tipo_pozo][wo_number] = {}

    if ruta_origen_workover:
        # Encuentra la primera subcarpeta que comienza con "2."
        ruta_primera_subcarpeta = encontrar_carpeta_por_prefijo(ruta_origen_workover, '2.')
        if ruta_primera_subcarpeta:
            # Encuentra la segunda subcarpeta que comienza con "6." dentro de la primera subcarpeta
            ruta_segunda_subcarpeta = encontrar_carpeta_por_prefijo(ruta_primera_subcarpeta, '6.')
            if ruta_segunda_subcarpeta:
                for archivo in os.listdir(ruta_segunda_subcarpeta):
                    if (archivo.endswith('.xlsx') or archivo.endswith('.xls') or archivo.endswith('.XLSX') or archivo.endswith('.XLS')) and "sumario" in archivo.lower():
                        shutil.copy(os.path.join(ruta_segunda_subcarpeta, archivo), ruta_destino)
                        estado_documento = 'OK'
                        break
    
    # Actualizar el estado del documento en el diccionario para el pozo correspondiente
    estado_documentos[pozo][tipo_pozo][wo_number]['Sumario'] = estado_documento

# Caso en que se desea copiar un archivo excel (Sumarios Finales) para CPI
def copiar_sumario_cpi(pozo, ruta_destino):
    ruta_base_origen = obtener_ruta_origen(pozo)
    ruta_cpi = obtener_directorio_paraCPI(ruta_base_origen)

    estado_documento = 'X' # Estado por defecto
    if pozo not in estado_documentos:
        estado_documentos[pozo] = {'CPI': {}}

    if ruta_cpi:
        # Encuentra la primera subcarpeta que comienza con "2."
        ruta_primera_subcarpeta = encontrar_carpeta_por_prefijo(ruta_cpi, '2.')
        if ruta_primera_subcarpeta:
            # Encuentra la segunda subcarpeta que comienza con "6." dentro de la primera subcarpeta
            ruta_segunda_subcarpeta = encontrar_carpeta_por_prefijo(ruta_primera_subcarpeta, '6.')
            if ruta_segunda_subcarpeta:
                for archivo in os.listdir(ruta_segunda_subcarpeta):
                    if (archivo.endswith('.xlsx') or archivo.endswith('.xls') or archivo.endswith('.XLSX') or archivo.endswith('.XLS')) and "sumario" in archivo.lower() :
                        shutil.copy(os.path.join(ruta_segunda_subcarpeta, archivo), ruta_destino)
                        estado_documento = 'OK'
                        break
    
    # Actualizar el estado del documento en el diccionario para el pozo correspondiente
    estado_documentos[pozo]['CPI']['Sumario'] = estado_documento

# Caso en que se desea copiar archivos pdf de AIS desde OneDrive   
def copiar_AIS_Opex(pozo, año_inicio, wo_number, ruta_destino, tipo_pozo, abandono):
    años = [str(año_inicio), str(año_inicio -1)] # Se busca en el año de interés y en el anterior
    encontrado = False # Bandera para saber si se encontró el archivo

    estado_documento = 'X' # Estado por defecto
    if 'adecua' in abandono.lower(): 
        pozo2 = f'{pozo} ADECUACION'
    else: 
        pozo2 = pozo
    Wo_number = str(wo_number)
    if pozo2 not in estado_documentos:
        estado_documentos[pozo2] = {}
    if tipo_pozo not in estado_documentos[pozo2]:
        estado_documentos[pozo2][tipo_pozo] = {}
    if wo_number not in estado_documentos[pozo2][tipo_pozo]:
        estado_documentos[pozo2][tipo_pozo][wo_number] = {}

    for año in años:
        if encontrado: # Si ya se encontró el archivo, se sale del bucle
            break
    
        # Obtener el nombre de usuario del sistema operativo
        usuario = os.getlogin()
        ruta_origen_onedrive = rf'C:\Users\{usuario}\OneDrive - SLB\Actas de Inicio y Recepcion Operaciones'
        ruta_base_AIS = os.path.join(ruta_origen_onedrive, f'AIS OPR {año}')
        ruta_siguiente = encontrar_carpeta_por_prefijo(ruta_base_AIS, '4.')
        if 'adecua' in abandono.lower():
            ruta_especifica = obtener_directorio_ARS(ruta_siguiente, pozo)
        else:
            ruta_especifica = obtener_directorio_AIS(ruta_siguiente, pozo, Wo_number)

        
        if ruta_especifica:
            archivos = [archivo for archivo in os.listdir(ruta_especifica) if archivo.endswith('.pdf') or archivo.endswith('.PDF')]
        
            # Filtramos por 'consolidado' y 'signed' en el nombre del archivo, y ordenamos por la cantidad de 'signed' en el nombre
            archivos_filtrados = sorted(
                [archivo for archivo in archivos if 'consolidado' in archivo.lower() and 'signed' in archivo.lower()],
                key = lambda x: x.count('signed'), reverse=True
            )

            if archivos_filtrados:
                archivo = archivos_filtrados[0] # Tomamos el archivo con más 'signed' en el nombre
                shutil.copy(os.path.join(ruta_especifica, archivo), ruta_destino)
                estado_documento = 'OK'
                encontrado = True

    # Actualizar el estado del documento en el diccionario para el pozo correspondiente
    estado_documentos[pozo2][tipo_pozo][wo_number]['AIS'] = estado_documento
        
def copiar_AIS(pozo, año_inicio, wo_number, ruta_destino, tipo_pozo):
    años = [str(año_inicio), str(año_inicio -1)] # Se busca en el año de interés y en el anterior
    encontrado = False # Bandera para saber si se encontró el archivo

    estado_documento = 'X' # Estado por defecto
    if tipo_pozo == 'Capex':
        Wo_number = str(wo_number)
        if pozo not in estado_documentos:
            estado_documentos[pozo] = {}
        if tipo_pozo not in estado_documentos[pozo]:
            estado_documentos[pozo][tipo_pozo] = {}
        if wo_number not in estado_documentos[pozo][tipo_pozo]:
            estado_documentos[pozo][tipo_pozo][wo_number] = {}
    
    else:
        if pozo not in estado_documentos:
            estado_documentos[pozo] = {'CPI': {}} # Se crea el diccionario para el pozo si no existe

    for año in años:
        if encontrado: # Si ya se encontró el archivo, se sale del bucle
            break
    
        # Obtener el nombre de usuario del sistema operativo
        usuario = os.getlogin()
        ruta_origen_onedrive = rf'C:\Users\{usuario}\OneDrive - SLB\Actas de Inicio y Recepcion Operaciones'
        ruta_base_AIS = os.path.join(ruta_origen_onedrive, f'AIS OPR {año}')
    
        # Caso en que el pozo es del tipo Capex
        if tipo_pozo == 'Capex':
            ruta_siguiente = encontrar_carpeta_por_prefijo(ruta_base_AIS, '3.')
            ruta_especifica = obtener_directorio_AIS(ruta_siguiente, pozo, Wo_number)

        elif tipo_pozo == 'CPI':
            ruta_siguiente = encontrar_carpeta_por_prefijo(ruta_base_AIS, '2.')
            ruta_especifica = obtener_directorio_AIS_CPI(ruta_siguiente, pozo)
        
        if ruta_especifica:
            archivos = [archivo for archivo in os.listdir(ruta_especifica) if archivo.endswith('.pdf') or archivo.endswith('.PDF')]
        
            # Filtramos por 'consolidado' y 'signed' en el nombre del archivo, y ordenamos por la cantidad de 'signed' en el nombre
            archivos_filtrados = sorted(
                [archivo for archivo in archivos if 'consolidado' in archivo.lower() and 'signed' in archivo.lower()],
                key = lambda x: x.count('signed'), reverse=True
            )

            if archivos_filtrados:
                archivo = archivos_filtrados[0] # Tomamos el archivo con más 'signed' en el nombre
                shutil.copy(os.path.join(ruta_especifica, archivo), ruta_destino)
                estado_documento = 'OK'
                encontrado = True

    # Actualizar el estado del documento en el diccionario para el pozo correspondiente
    if tipo_pozo == 'Capex':
        estado_documentos[pozo][tipo_pozo][wo_number]['AIS'] = estado_documento
    else:
        estado_documentos[pozo]['CPI']['AIS'] = estado_documento

# Caso en que se desea copiar archivos de Diagramas Mecanicos
def copiar_diagramas_mecanicos(pozo, wo_number, ruta_destino, tipo_pozo, abandono):
    ruta_base_origen = obtener_ruta_origen(pozo)

    estado_documento = 'X' # Estado por defecto
    if 'adecua' in abandono.lower():
        pozo2 = f'{pozo} ADECUACION'
    else: 
        pozo2 = pozo
    if pozo2 not in estado_documentos:
        estado_documentos[pozo2] = {}
    if tipo_pozo not in estado_documentos[pozo2]:
        estado_documentos[pozo2][tipo_pozo] = {}
    if wo_number not in estado_documentos[pozo2][tipo_pozo]:
        estado_documentos[pozo2][tipo_pozo][wo_number] = {}

    if ruta_base_origen:
        # Encuentra la subcarpeta que contiene Diagramas en su nombre 
        for carpeta in os.listdir(ruta_base_origen):
            if "DIAGRAMAS" in carpeta.upper():
                ruta_diagramas = os.path.join(ruta_base_origen, carpeta)
                # Utilizamos la expresión regular para encontrar archivos relevantes
                if "adecua" in abandono.lower():
                    patron = re.compile(rf'.*{pozo}.*ABANDONO.*', re.IGNORECASE)
                else:
                    patron = re.compile(rf'.*{pozo}.*WO.*{wo_number}.*', re.IGNORECASE)
                
                for archivo in os.listdir(ruta_diagramas):
                    if patron.match(archivo):
                        shutil.copy(os.path.join(ruta_diagramas, archivo), ruta_destino)
                        estado_documento = 'OK'
                        # El bucle no se rompe para copiar todos los archivos que coincidan
                break
    
    # Actualizar el estado del documento en el diccionario para el pozo correspondiente
    estado_documentos[pozo2][tipo_pozo][wo_number]['Diagramas Mecánicos'] = estado_documento

# Caso en que se desea copiar archivos de registro 
def copiar_registros(pozo, wo_number, ruta_destino, tipo_pozo, abandono):
    ruta_base_origen = obtener_ruta_origen(pozo)
    ruta_origen_workover = obtener_directorio_por_workover(ruta_base_origen, wo_number)

    estado_documento = 'X' # Estado por defecto
    if 'adecua' in abandono.lower():
        pozo2 = f'{pozo} ADECUACION'
    else:
        pozo2 = pozo
    if pozo2 not in estado_documentos:
        estado_documentos[pozo2] = {}
    if tipo_pozo not in estado_documentos[pozo2]:
        estado_documentos[pozo2][tipo_pozo] = {}
    if wo_number not in estado_documentos[pozo2][tipo_pozo]:
        estado_documentos[pozo2][tipo_pozo][wo_number] = {}

    if ruta_origen_workover:
        # Encuentra la primera subcarpeta que comienza con "2."
        ruta_primera_subcarpeta = encontrar_carpeta_por_prefijo(ruta_origen_workover, '2.')
        if ruta_primera_subcarpeta:
            # Busca dentro de "2. Ejecución" la carpeta que contiene "Log" en su nombre
            for nombre in os.listdir(ruta_primera_subcarpeta):
                if "LOG" in nombre.upper():
                    ruta_segunda_subcarpeta = os.path.join(ruta_primera_subcarpeta, nombre)
                    # Utilizamos la expresión regular para encontrar archivos relevantes
                    #if "adecua" in abandono.lower():
                    #    patron = re.compile(rf'.*{pozo}.*ABANDONO.*', re.IGNORECASE)
                    #else:
                    #    patron = re.compile(rf'.*{pozo}.*WO.*{wo_number}.*', re.IGNORECASE)
                    patron = re.compile(rf'.*{pozo}.*WO.*{wo_number}.*', re.IGNORECASE)

                    for archivo in os.listdir(ruta_segunda_subcarpeta):
                        if patron.match(archivo):
                            shutil.copy(os.path.join(ruta_segunda_subcarpeta, archivo), ruta_destino)
                            estado_documento = 'OK'
                            # El bucle no se rompe para copiar todos los archivos que coincidan
                    break

    # Actualizar el estado del documento en el diccionario para el pozo correspondiente            
    estado_documentos[pozo2][tipo_pozo][wo_number]['Registros'] = estado_documento
    

# Caso en que se desea copiar un archivo pptx de Post Mortem desde OneDrive, caso Opex (y caso especial abandono, tipo de Opex)
def copiar_PM_Opex(pozo, año_fin, wo_number, ruta_destino, tipo_pozo, abandono):
    años = [str(año_fin), str(año_fin + 1)] # Se busca en el año de interés y en el siguiente
    encontrado = False # Bander para saber si se encontró el archvio 
    Wo_number = str(wo_number)

    estado_documento = 'X' # Estado por defecto
    if 'adecua' in abandono.lower():
        pozo2 = f'{pozo} ADECUACION'
    else:
        pozo2 = pozo
    if pozo2 not in estado_documentos:
        estado_documentos[pozo2] = {}
    if tipo_pozo not in estado_documentos[pozo2]:
        estado_documentos[pozo2][tipo_pozo] = {}
    if wo_number not in estado_documentos[pozo2][tipo_pozo]:
        estado_documentos[pozo2][tipo_pozo][wo_number] = {}

    # Compilar patrones de búsqueda para reutilizar en el filtrado
    patron_opex = re.compile(rf'.*{pozo}.*{Wo_number}.*OK', re.IGNORECASE)
    patron_abandono = re.compile(rf'.*{pozo}.*OK', re.IGNORECASE)

    for año in años:
        if encontrado: # Si ya se encontró el archivo, se sale del bucle
            break
    
        # Obtener el nombre de usuario del sistema operativo
        usuario = os.getlogin()
        ruta_origen_onedrive = rf'C:\Users\{usuario}\OneDrive - SLB\General - Post mortem\{año}'
        
        # Verificar si la carpeta del año existe
        if os.path.exists(ruta_origen_onedrive):
            for carpeta in os.listdir(ruta_origen_onedrive):
                ruta_carpeta_mes = os.path.join(ruta_origen_onedrive, carpeta)

                if os.path.isdir(ruta_carpeta_mes):
                    # Buscar archivos pptx en la carpeta
                    archivos = [archivo for archivo in os.listdir(ruta_carpeta_mes) if archivo.endswith('.pptx') or archivo.endswith('.PPTX')]

                    for archivo in archivos:
                        # Filtrar por el patrón de búsqueda
                        archivo_filtrado = [archivo for archivo in archivos if patron_opex.search(archivo)]

                        # Copiar el archivo encontrado
                        if archivo_filtrado:
                            shutil.copy(os.path.join(ruta_carpeta_mes, archivo_filtrado[0]), ruta_destino)
                            estado_documento = 'OK'
                            encontrado = True
                            break
                
                if encontrado:
                    break
    
    # Actualizar el estado del documento en el diccionario para el pozo correspondiente      
    estado_documentos[pozo2][tipo_pozo][wo_number]['Post Mortem'] = estado_documento
    
# Caso en que se desea copiar archivos pdf de ARS desde OneDrive
def copiar_ARS_Opex(pozo, año_fin, wo_number, ruta_destino, tipo_pozo, abandono):
    años = [str(año_fin), str(año_fin + 1)] # Se busca en el año de interés y en el siguiente
    encontrado = False # Bander para saber si se encontró el archvio 
    
    estado_documento = 'X' # Estado por defecto
    Wo_number = str(wo_number)
    if 'adecua' in abandono.lower():
        pozo2 = f'{pozo} ADECUACION'
    else:
        pozo2 = pozo
    if pozo2 not in estado_documentos:
        estado_documentos[pozo2] = {}
    if tipo_pozo not in estado_documentos[pozo2]:
        estado_documentos[pozo2][tipo_pozo] = {}
    if wo_number not in estado_documentos[pozo2][tipo_pozo]:
        estado_documentos[pozo2][tipo_pozo][wo_number] = {}

    for año in años:
        if encontrado: # Si ya se encontró el archivo, se sale del bucle
            break
    
        # Obtener el nombre de usuario del sistema operativo
        usuario = os.getlogin()
        ruta_origen_onedrive = rf'C:\Users\{usuario}\OneDrive - SLB\Actas de Inicio y Recepcion Operaciones'
        ruta_base_ARS = os.path.join(ruta_origen_onedrive, f'ACTA DE RECEPCION DE OPERACIONES {año}')
    
        if 'adecua' in abandono.lower():
            ruta_siguiente = obtener_directorio_ARS(ruta_base_ARS, pozo)
        else:
            ruta_siguiente = obtener_directorio_AIS(ruta_base_ARS, pozo, Wo_number)

        if ruta_siguiente:
            archivos = [archivo for archivo in os.listdir(ruta_siguiente) if archivo.endswith('.pdf') or archivo.endswith('.PDF')]
        
            # Filtramos por 'consolidado' y 'signed' en el nombre del archivo, y ordenamos por la cantidad de 'signed' en el nombre
            archivos_filtrados = sorted(
                [archivo for archivo in archivos if 'consolidado' in archivo.lower() and 'signed' in archivo.lower()],
                key = lambda x: x.count('signed'), reverse=True
            )

            if archivos_filtrados:
                archivo = archivos_filtrados[0] # Tomamos el archivo con más 'signed' en el nombre
                shutil.copy(os.path.join(ruta_siguiente, archivo), ruta_destino)
                estado_documento = 'OK'
                encontrado = True

    # Actualizar el estado del documento en el diccionario para el pozo correspondiente
    estado_documentos[pozo2][tipo_pozo][wo_number]['ARS'] = estado_documento
    
def copiar_ARS(pozo, año_fin, wo_number, ruta_destino, tipo_pozo):
    años = [str(año_fin), str(año_fin + 1)] # Se busca en el año de interés y en el siguiente
    encontrado = False # Bander para saber si se encontró el archvio 

    estado_documento = 'X' # Estado por defecto
    if pozo not in estado_documentos:
        estado_documentos[pozo] = {}
    if tipo_pozo not in estado_documentos[pozo]:
        estado_documentos[pozo][tipo_pozo] = {}
    if wo_number not in estado_documentos[pozo][tipo_pozo]:
        estado_documentos[pozo][tipo_pozo][wo_number] = {}

    for año in años:
        if encontrado: # Si ya se encontró el archivo, se sale del bucle
            break
    
        # Obtener el nombre de usuario del sistema operativo
        usuario = os.getlogin()
        ruta_origen_onedrive = rf'C:\Users\{usuario}\OneDrive - SLB\Actas de Inicio y Recepcion Operaciones'
        ruta_base_ARS = os.path.join(ruta_origen_onedrive, f'ACTA DE RECEPCION DE OPERACIONES {año}')
    
        if tipo_pozo == 'Capex':
            Wo_number = str(wo_number)
            ruta_siguiente = obtener_directorio_AIS(ruta_base_ARS, pozo, Wo_number)
        
        elif tipo_pozo == 'CPI':
            ruta_siguiente = obtener_directorio_AIS_CPI(ruta_base_ARS, pozo)
        

        if ruta_siguiente:
            
            archivos = [archivo for archivo in os.listdir(ruta_siguiente) if archivo.endswith('.pdf') or archivo.endswith('.PDF')]
        
            # Filtramos por 'consolidado' y 'signed' en el nombre del archivo, y ordenamos por la cantidad de 'signed' en el nombre
            archivos_filtrados = sorted(
                [archivo for archivo in archivos if 'consolidado' in archivo.lower() and 'signed' in archivo.lower()],
                key = lambda x: x.count('signed'), reverse=True
            )

            if archivos_filtrados:
                archivo = archivos_filtrados[0] # Tomamos el archivo con más 'signed' en el nombre
                shutil.copy(os.path.join(ruta_siguiente, archivo), ruta_destino)
                estado_documento = 'OK'
                encontrado = True

    # Actualizar el estado del documento en el diccionario para el pozo correspondiente
    if tipo_pozo == 'Capex':
        estado_documentos[pozo][tipo_pozo][wo_number]['ARS'] = estado_documento
    else:
        estado_documentos[pozo]['CPI']['ARS'] = estado_documento

# Función para crear directorios
def crear_directorios(df, ruta_trimestre, tipo_actividad):
    os.makedirs(ruta_trimestre, exist_ok=True)

    if tipo_actividad == 'CPI': # Caso en que se desea crear directorios para CPI
        ruta_completacion = os.path.join(ruta_trimestre, 'CPI')
        os.makedirs(ruta_completacion, exist_ok=True)
        for _, fila in df.iterrows(): # De cada pozo, se obtiene su información para buscar sus archivos y crear los directorios del reporte
            año_inicio = fila['START_WO'].year
            año_fin = fila['END_WO'].year
            pozo = fila['ITEM_NAME']
            abandono = ""
            nombre_carpeta_pozo = f'{pozo}'
            #crear carpeta del pozo
            ruta_carpeta_pozo = os.path.join(ruta_completacion, nombre_carpeta_pozo)
            os.makedirs(ruta_carpeta_pozo, exist_ok=True)

            # Crear subcarpetas dentro de la carpeta del pozo
            subcarpetas = ['1.1 Prognosis y Programas', '1.2 Soporte para inicio de Trabajos', 
                       '1.3 Reportes diarios de ejecución', '1.4 Sumarios Finales CPI', 
                       '1.5 Actas de Inicio y Recepción']
            for subcarpeta in subcarpetas:
                os.makedirs(os.path.join(ruta_carpeta_pozo, subcarpeta), exist_ok=True)

            # Llamar a copiar_prognosis_cpi para copiar la Propuesta Técnica respectiva en la subcarpeta 1.1 Prognosis y Programas
            ruta_destino_prognosis = os.path.join(ruta_carpeta_pozo, '1.1 Prognosis y Programas')
            copiar_prognosis_cpi(pozo, ruta_destino_prognosis)

            # Llamar a copiar_pdf_cpi para copiar la Notificación respectiva en la subcarpeta 1.2 Soporte para inicio de Trabajos
            ruta_destino_soporte = os.path.join(ruta_carpeta_pozo, '1.2 Soporte para inicio de Trabajos')
            copiar_pdf_cpi(pozo, ruta_destino_soporte)

            # Llamar a copiar_reportes_diarios_cpi para copiar los Reportes Diarios respectivos en la subcarpeta 1.3 Reportes diarios de ejecución
            ruta_destino_reportes = os.path.join(ruta_carpeta_pozo, '1.3 Reportes diarios de ejecución')
            copiar_reportes_diarios_cpi(pozo, ruta_destino_reportes)

            # Llamar a copiar_sumario_cpi para copiar el Sumario respectivo en la subcarpeta 1.4 Sumarios Finales CPI
            ruta_destino_sumario = os.path.join(ruta_carpeta_pozo, '1.4 Sumarios Finales CPI')
            copiar_sumario_cpi(pozo, ruta_destino_sumario)

            # Llamar a copiar_AIS para copiar el AIS respectivo en la subcarpeta 1.5 Actas de Inicio y Recepción
            ruta_destino_AIS = os.path.join(ruta_carpeta_pozo, '1.5 Actas de Inicio y Recepción', 'AIS')
            os.makedirs(ruta_destino_AIS, exist_ok=True)
            copiar_AIS(pozo, año_inicio, None, ruta_destino_AIS, 'CPI')

            # Llamar a copiar_ARS para copiar el ARS respectivo en la subcarpeta 1.5 Actas de Inicio y Recepción
            ruta_destino_ARS = os.path.join(ruta_carpeta_pozo, '1.5 Actas de Inicio y Recepción', 'ARS')
            os.makedirs(ruta_destino_ARS, exist_ok=True)
            copiar_ARS(pozo, año_fin, None, ruta_destino_ARS, 'CPI')

    elif tipo_actividad == 'CAPEX': # Caso en que se desea crear directorios para Capex
        for _, fila in df.iterrows(): # De cada pozo, se obtiene su información para buscar sus archivos y crear los directorios del reporte
            año_inicio = fila['START_WO'].year
            año_fin = fila['END_WO'].year
            pozo = fila['ITEM_NAME']
            wo_number = int(fila['WO_NUMBER'])
            abandono = ""
            nombre_carpeta_pozo = f'{pozo} WO {wo_number}'
            ruta_carpeta_pozo = os.path.join(ruta_trimestre, nombre_carpeta_pozo)
            os.makedirs(ruta_carpeta_pozo, exist_ok=True)

            # Crear subcarpetas dentro de la carpeta del pozo
            subcarpetas = ['2.1 Prognosis y Programas', '2.2 Soporte para inicio de Trabajos', 
                       '2.3 Reportes diarios de ejecución',  '2.4 Sumarios Finales WO', 
                       '2.5 Actas de Inicio y Recepción']
            for subcarpeta in subcarpetas:
                ruta_subcarpeta = os.path.join(ruta_carpeta_pozo, subcarpeta)
                os.makedirs(ruta_subcarpeta, exist_ok=True)

            # Llamar a copiar_archivo_prognosis para copiar la Propuesta Técnica respectiva en la subcarpeta 2.1 Prognosis y Programas
            ruta_destino_prognosis = os.path.join(ruta_carpeta_pozo, '2.1 Prognosis y Programas')
            copiar_prognosis_capex(pozo, int(wo_number), ruta_destino_prognosis, "Capex", abandono)

            # Llamar a copiar_pdf para copiar la Notificación respectiva en la subcarpeta 2.2 Soporte para inicio de Trabajos
            ruta_destino_soporte = os.path.join(ruta_carpeta_pozo, '2.2 Soporte para inicio de Trabajos')
            copiar_pdf(pozo, int(wo_number), ruta_destino_soporte, "Capex", abandono)

            # Llamar a copiar_reportes_diarios para copiar los Reportes Diarios respectivos en la subcarpeta 2.3 Reportes diarios de ejecución
            ruta_destino_reportes = os.path.join(ruta_carpeta_pozo, '2.3 Reportes diarios de ejecución')
            copiar_reportes_diarios(pozo, int(wo_number), ruta_destino_reportes, "Capex", abandono)

            # Llamar a copiar_sumario para copiar el Sumario respectivo en la subcarpeta 2.4 Sumarios Finales WO
            ruta_destino_sumario = os.path.join(ruta_carpeta_pozo, '2.4 Sumarios Finales WO')
            copiar_sumario(pozo, int(wo_number), ruta_destino_sumario, "Capex", abandono)

            # Llamar a copiar_AIS para copiar el AIS respectivo en la subcarpeta 2.5 Actas de Inicio y Recepción
            ruta_destino_AIS = os.path.join(ruta_carpeta_pozo, '2.5 Actas de Inicio y Recepción', 'AIS')
            os.makedirs(ruta_destino_AIS, exist_ok=True)
            copiar_AIS(pozo, año_inicio, int(wo_number), ruta_destino_AIS, 'Capex')

            # Llamar a copiar_ARS para copiar el ARS respectivo en la subcarpeta 2.5 Actas de Inicio y Recepción
            ruta_destino_ARS = os.path.join(ruta_carpeta_pozo, '2.5 Actas de Inicio y Recepción', 'ARS')
            os.makedirs(ruta_destino_ARS, exist_ok=True)
            copiar_ARS(pozo, año_fin, int(wo_number), ruta_destino_ARS, 'Capex')

    elif tipo_actividad == 'OPEX': # Caso en que se desea crear directorios para Opex
        for _, fila in df.iterrows(): # De cada pozo, se obtiene su información para buscar sus archivos y crear los directorios del reporte
            año_inicio = fila['START_WO'].year
            año_fin = fila['END_WO'].year
            pozo = fila['ITEM_NAME']
            wo_number = fila['WO_NUMBER'] if pd.notna(fila['WO_NUMBER']) else None
            abandono = fila['WO_OPEX_TEXT'] if pd.notna(fila['WO_OPEX_TEXT']) else ""

            if 'adecua' in abandono.lower(): # Si el pozo es de adecuación, se cambia el nombre de la carpeta para conveniencia
                nombre_carpeta_pozo = f'{pozo} {int(wo_number)} ADECUACION'
            else:
                nombre_carpeta_pozo = f'{pozo} WO {int(wo_number)}'
            ruta_carpeta_pozo = os.path.join(ruta_trimestre, nombre_carpeta_pozo)
            os.makedirs(ruta_carpeta_pozo, exist_ok=True)

            # Crear subcarpetas dentro de la carpeta del pozo
            subcarpetas = ['3.1 Prognosis y Programas', '3.2 Soporte para inicio de Trabajos', 
                           '3.3 Reportes diarios de ejecución', '3.4 Diagramas Mecánicos', 
                           '3.5 Registros', '3.6 Sumarios Finales', '3.7 Post Mortem Pozos', 
                           '3.8 Actas de Inicio y Recepción']
            
            # Crear las subcarpetas excepto '3.5 Registros'
            for subcarpeta in subcarpetas: 
                if subcarpeta != '3.5 Registros':
                    ruta_subcarpeta = os.path.join(ruta_carpeta_pozo, subcarpeta)
                    os.makedirs(ruta_subcarpeta, exist_ok=True)
                
            # Llamar a copiar_archivo_prognosis para copiar la Propuesta Técnica respectiva en la subcarpeta 3.1 Prognosis y Programas
            ruta_destino_prognosis = os.path.join(ruta_carpeta_pozo, '3.1 Prognosis y Programas')
            copiar_prognosis_capex(pozo, int(wo_number), ruta_destino_prognosis, "OPEX", abandono)

            # Llamar a copiar_pdf para copiar la Notificación respectiva en la subcarpeta 3.2 Soporte para inicio de Trabajos
            ruta_destino_soporte = os.path.join(ruta_carpeta_pozo, '3.2 Soporte para inicio de Trabajos')
            copiar_pdf(pozo, int(wo_number), ruta_destino_soporte, "OPEX", abandono)

            # Llamar a copiar_reportes_diarios para copiar los Reportes Diarios respectivos en la subcarpeta 3.3 Reportes diarios de ejecución
            ruta_destino_reportes = os.path.join(ruta_carpeta_pozo, '3.3 Reportes diarios de ejecución')
            copiar_reportes_diarios(pozo, int(wo_number), ruta_destino_reportes, "OPEX", abandono)

            # Llamar a copiar_sumario para copiar el Sumario respectivo en la subcarpeta 3.6 Sumarios Finales
            ruta_destino_sumario = os.path.join(ruta_carpeta_pozo, '3.6 Sumarios Finales')
            copiar_sumario(pozo, int(wo_number), ruta_destino_sumario, "OPEX", abandono)

            # Llamar a copiar_AIS para copiar el AIS respectivo en la subcarpeta 3.8 Actas de Inicio y Recepción
            ruta_destino_AIS = os.path.join(ruta_carpeta_pozo, '3.8 Actas de Inicio y Recepción', 'AIS')
            os.makedirs(ruta_destino_AIS, exist_ok=True)
            copiar_AIS_Opex(pozo, año_inicio, int(wo_number), ruta_destino_AIS, 'OPEX', abandono)

            # Llamar a copiar_ARS para copiar el ARS respectivo en la subcarpeta 3.8 Actas de Inicio y Recepción
            ruta_destino_ARS = os.path.join(ruta_carpeta_pozo, '3.8 Actas de Inicio y Recepción', 'ARS')
            os.makedirs(ruta_destino_ARS, exist_ok=True)
            copiar_ARS_Opex(pozo, año_fin, int(wo_number), ruta_destino_ARS, 'OPEX', abandono)

            # Llamar a copiar_pm_Opex para copiar la Presentación/ Post Mortem del pozo respectivo en la subcarpeta 3.7 Post Mortem Pozos
            ruta_destino_pm = os.path.join(ruta_carpeta_pozo, '3.7 Post Mortem Pozos')
            os.makedirs(ruta_destino_pm, exist_ok=True)
            copiar_PM_Opex(pozo, año_fin, int(wo_number), ruta_destino_pm, 'OPEX', abandono)

            # Llamar a copiar_diagramas_mecanicos para copiar el Diagrama respectivo en la subcarpeta 3.4 Diagramas Mecánicos
            ruta_destino_diagramas = os.path.join(ruta_carpeta_pozo, '3.4 Diagramas Mecánicos')
            os.makedirs(ruta_destino_diagramas, exist_ok=True)
            copiar_diagramas_mecanicos(pozo, int(wo_number), ruta_destino_diagramas, "OPEX", abandono)

            # Verificacion adicional para '3.5 Registros'   
            # De cada pozo con actividad Opex, solamente si contiene la carpeta "Log", sabemos que tendrá registros.

            # Primero buscamos la carpeta del Pozo en el directorio 00_Well_File
            # De cada pozo con actividad Opex, solamente si contiene la carpeta "Log", sabemos que tendrá registros.
            # Una vez establecido si la tiene o no, se creará carpeta de Registros en el Reporte y se copiarán los registros respectivos
            ruta_base_origen = obtener_ruta_origen(pozo)

            # Esta funcion devuelve la ruta completa o None si no se encuentra nada
            ruta_origen_workover = obtener_directorio_por_workover(ruta_base_origen, int(wo_number))
            if ruta_origen_workover:
                # Encuentra la primera subcarpeta que comienza con "2."
                ruta_primera_subcarpeta = encontrar_carpeta_por_prefijo(ruta_origen_workover, '2.')
                if ruta_primera_subcarpeta:
                    # Busca dentro de "2. Ejecución" la carpeta que contiene "Log" en su nombre
                    for nombre in os.listdir(ruta_primera_subcarpeta):
                        if "LOG" in nombre.upper():
                            ruta_destino_registros = os.path.join(ruta_carpeta_pozo, '3.5 Registros')
                            os.makedirs(ruta_destino_registros, exist_ok=True)
                            copiar_registros(pozo, int(wo_number), ruta_destino_registros, "OPEX", abandono)
                        else:
                            if 'adecua' in abandono.lower():
                                pozo2 = f'{pozo} ADECUACION'
                                estado_documentos[pozo2][tipo_actividad][wo_number]['Registros'] = 'N/A'
                            else:
                                estado_documentos[pozo][tipo_actividad][wo_number]['Registros'] = 'N/A'

    


def calcular_porcentaje_completitud(row):
    campos_relevantes = ['Propuesta Técnica', 'Notificación', 'Reportes Diarios', 'Sumario', 'AIS', 'ARS', 'Post Mortem', 'Diagramas Mecánicos', 'Registros']
    total_campos = 0
    campos_completados = 0

    for campo in campos_relevantes:
        # Nos aseguramos de que sólo se cuenten los campos con datos
        estado = row.get(campo, "")
        if estado != "" and estado != "N/A" and estado is not None:
            total_campos += 1
            if "OK" in estado: 
                campos_completados += 1

    # Se calcula el porcentaje si hay campos a considerar
    if total_campos > 0:
        porcentaje = (campos_completados / total_campos) * 100
        return f"{porcentaje:.2f}%" # Se redondea a dos decimales
    else:
        return "N/A" # Si no hay campos a considerar, se retorna N/A

def generar_informe_excel(estado_documentos, año_interes, trimestre_actual, directorio_base):
    
    # se crea un dataframe a partir del diccionario
    data = []
    
    # Se recorre el diccionario para obtener los datos
    for pozo, pozo_info in estado_documentos.items():
        for tipo, tipo_info in pozo_info.items():
            # Se separa el caso de CPI, ya que no tiene WO #
            if tipo == 'CPI':
                cpi_info = {k: v for k, v in tipo_info.items() if k is not None}
                data.append([pozo, tipo, None] + list(cpi_info.values())) 
            else:
                for wo_number, values in tipo_info.items():
                    row = [pozo, tipo, wo_number] + list(values.values())
                    data.append(row)
                    
    # se crea el dataframe
    columns = ["Pozo", "Tipo", "WO #", "Propuesta Técnica", "Notificación", "Reportes Diarios", "Sumario", "AIS", "ARS", "Post Mortem", 'Diagramas Mecánicos', "Registros"]
    df = pd.DataFrame(data, columns=columns)
    df.insert(len(df.columns), "Apoyo a la Operación", "OK, Para Pestaña de Control Documental")

    # Nombre del archivo basado en el trimestre y año actuales
    hora_actual = now.strftime("%H%M%S")
    nombre_archivo = f'Informe Trimestral Q{trimestre_actual} {año_interes} ({hora_actual}).xlsx'
    ruta_guardado = os.path.join(directorio_base, nombre_archivo)

    for archivo in os.listdir(directorio_base):
        if archivo.startswith(f'Informe Trimestral Q{trimestre_actual} {año_interes}') and archivo.endswith('.xlsx'):
            os.remove(os.path.join(directorio_base, archivo))

    df.fillna("N/A", inplace=True)  # Reemplazar valores nulos por "N/A" (Not Available)
    

    # Calcular el porcentaje de completitud y añadirlo como una columna al dataframe
    df['Porcentaje de Avance'] = df.apply(calcular_porcentaje_completitud, axis=1)

    # Guardar en un archivo excel
    with pd.ExcelWriter(ruta_guardado, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Ing. Prod-Eje.', index=False, startrow=1)
        
        # Obtenemos la hoja de trabajo de openpyxl para poder editarla
        workbook = writer.book
        worksheet = writer.sheets['Ing. Prod-Eje.']

        # Aplicar formato a las celdas con "X" y "OK, Para Pestaña de Control Documental"
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        light_blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        for row in worksheet.iter_rows(min_row=3, max_col=len(df.columns), max_row=worksheet.max_row):
            for cell in row:
                if cell.value == "X":
                    cell.fill = red_fill
                elif cell.value == "OK, Para Pestaña de Control Documental":
                    cell.fill = light_blue_fill

        # Calcular y añadir el total de los porcentajes de completitud al final del documento
        
        worksheet.append([])
        total_row = len(df) + 4
        total_percentage = sum(df['Porcentaje de Avance'].str.rstrip('%').astype(float)) / len(df)
        total_cell = worksheet.cell(row=total_row, column=len(df.columns) - 1)
        total_cell.value = "Total de Avance"
        total_cell.font = Font(bold=True)
        total_percentage_cell = worksheet.cell(row=total_row, column=len(df.columns))
        total_percentage_cell.value = f"{total_percentage:.2f}%"
        total_percentage_cell.font = Font(bold=True)

        
        # Aplicar formato a las celdas
        for col in worksheet.columns:
            max_length = max((len(str(cell.value)) for cell in col)) + 2
            worksheet.column_dimensions[col[0].column_letter].width = max_length
        
        # Aplicar negrita a los títulos de las columnas y centrar todo el texto
        for cell in worksheet["2:2"]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Centrar todas las celdas
        for row in worksheet.iter_rows(min_row=1, max_col=len(df.columns), max_row=len(df)+2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
        
        # Agregar título sobre los nombres de las columnas
        worksheet.insert_rows(1)
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = f'Informe Trimestral Q{trimestre_actual} {año_interes}'
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        # Combinar las celdas para el título
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

# Conexión a la base de datos y obtención de los DataFrames
conn = pyodbc.connect(conn_str)

# Consultas SQL
queryO = f""" 
SELECT VENDOR_ID_TEXT, ITEM_NAME, START_WO, END_WO, ESTIMATED_DATE, PLAN_TYPE_TEXT, WO_NUMBER, WO_OPEX_TEXT
FROM VT_WELLJOBLOG_en_US
WHERE YEAR(END_WO) = {año_interes} AND PLAN_TYPE_TEXT = 'Opex' AND VERIFIED = 'True'
"""

queryC = f""" 
SELECT VENDOR_ID_TEXT, ITEM_NAME, START_WO, END_WO, ESTIMATED_DATE, WO_NUMBER
FROM VT_WELLJOBLOG_en_US
WHERE YEAR(END_WO) = {año_interes} AND PLAN_TYPE_TEXT = 'Capex' AND VERIFIED = 'True'
"""

queryCompletacion = f"""
SELECT VENDOR_ID_TEXT, ITEM_NAME, START_WO, END_WO, ESTIMATED_DATE, WELL_STATUS_TEXT, WO_NUMBER
FROM VT_WELLJOBLOG_en_US
WHERE YEAR(END_WO) = {año_interes} AND WELL_STATUS_TEXT = 'CPI' AND VERIFIED = 'True'
"""

dfO = obtener_dataframe(queryO, conn)
dfC = obtener_dataframe(queryC, conn)
dfCo = obtener_dataframe(queryCompletacion, conn)
conn.close()

# Directorio base
# obtener automaticamente el usuario 
usuario = os.getlogin()

# Donde se creará el reporte trimestral
d = rf'\\dir.slb.com\NSA\SAM_Collaborate\EC0037\90_Public\01 INFORMES TRIMESTRALES'
directorio_base = os.path.join(d, f'Reporte Trimestral {año_interes}')

# Donde se creará el archivo excel de Seguimiento del Reporte Trimestral
directorio_base_Apoyo = rf'C:\Users\{usuario}\OneDrive - SLB\General - DIGITALIZACION ING. PROD. & EJEC_\01 Seguimiento Informe Trimestral'

os.makedirs(directorio_base, exist_ok=True)

# Filtrar DataFrames por trimestre
trimestres = [(1, 2, 3), (4, 5, 6), (7, 8, 9), (10, 11, 12)]
nombres_trimestres = [f'Reporte Trimestral Q1 {año_interes}', f'Reporte Trimestral Q2 {año_interes}', 
                      f'Reporte Trimestral Q3 {año_interes}', f'Reporte Trimestral Q4 {año_interes}']

# Solo procesamos el trimestre actual
meses_trimestre = trimestres[trimestre_actual - 1]
nombre_trimestre_actual = nombres_trimestres[trimestre_actual - 1]

# Filtramos los DataFrames por el trimestre actual
dfO_trimestre = filtrar_por_trimestreCPI_Opex(dfO, meses_trimestre)
dfC_trimestre = filtrar_por_trimestreCapex(dfC, meses_trimestre)
dfCo_trimestre = filtrar_por_trimestreCPI_Opex(dfCo, meses_trimestre)

# Dataframe combinado que representa todas las actividades de los pozos en el trimestre
df_combinado = pd.concat([dfO_trimestre, dfC_trimestre, dfCo_trimestre])

# Agregamos una columna que enumere los pozos
df_combinado.reset_index(drop=True, inplace=True)
df_combinado.index += 1 # Comenzar la enumeración desde 1

# Calcular los días programados y reales
df_combinado['Días Programados'] = (df_combinado['ESTIMATED_DATE'] - df_combinado['START_WO']).dt.days
df_combinado['Días Reales'] = (df_combinado['END_WO'] - df_combinado['START_WO']).dt.days

# Crear la columna de Comparación de Resultados y Problemas Operacionales
df_combinado['Comparación de Resultados'] = 'N/A'
df_combinado['Problemas Operacionales'] = 'Ver Openwells'

# Selecionar y renombar las columnas
df_final = df_combinado.rename(columns={
        'VENDOR_ID_TEXT': 'Taladro',
        'ITEM_NAME': 'Pozo',
        'START_WO': 'Fecha de Inicio',
        'END_WO': 'Fecha de Terminación',
    }).loc[:, ['Taladro', 'Pozo', 'Fecha de Inicio', 'Fecha de Terminación',
                'Días Programados', 'Días Reales', 'Comparación de Resultados', 'Problemas Operacionales']]
    
# Insertar la columna de numeración de items
df_final.insert(0, 'Item', df_final.index)

# Crear directorio para Apoyo a la Operación

# Aquí se crea la carpeta que va en el reporte trimestral
ruta_apoyo_operacion = os.path.join(directorio_base, nombre_trimestre_actual, '4. Apoyo a la Operación')

# Aquí vamos a guardar el archivo Excel de Apoyo a la Operación lleno con la pestaña de Control Documental
ruta_documento_AO = rf'\\dir.slb.com\NSA\SAM_Collaborate\EC0037\10_Development_execution\APOYO A LA OPERACIÓN'
os.makedirs(ruta_apoyo_operacion, exist_ok=True)

# Crear un archivo Excel con la pestaña correspondiente a Control Documental en la carpeta reciéntemente creada
nombre_archivo_excel = f"Apoyo Operación Q{trimestre_actual} {año_interes}.xlsx"
ruta_archivo_excel = os.path.join(ruta_documento_AO, nombre_archivo_excel)

# Guardar el DataFrame en el archivo Excel
with pd.ExcelWriter(ruta_archivo_excel, engine='openpyxl') as writer:
    df_final.to_excel(writer, sheet_name='Intervenciones con Torre', index=False, startrow=1)
        
    # Obtenemos la hoja de trabajo de openpyxl para poder editarla
    workbook = writer.book
    worksheet = writer.sheets['Intervenciones con Torre']
        
    # Aplicar formato a las celdas
    for col in worksheet.columns:
        max_length = max((len(str(cell.value)) for cell in col)) + 2
        worksheet.column_dimensions[col[0].column_letter].width = max_length
        
    # Aplicar negrita a los títulos de las columnas y centrar todo el texto
    for cell in worksheet["2:2"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Centrar todas las celdas
    for row in worksheet.iter_rows(min_row=1, max_col=len(df_final.columns), max_row=len(df_final)+2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
        
    # Agregar título sobre los nombres de las columnas
    worksheet.insert_rows(1)
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = f'Shaya Ecuador - Intervenciones con Torre  Q{trimestre_actual} {año_interes}'
    title_cell.font = Font(bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # Combinar las celdas para el título
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_final.columns))

# Crear directorios para cada pozo
ruta_trimestre_CPI = os.path.join(directorio_base, nombre_trimestre_actual, '1. Perforación & Completación de Pozos Nuevos')
ruta_trimestre_Opex = os.path.join(directorio_base, nombre_trimestre_actual, '3. Intervenciones con Torre (Pulling)')
ruta_trimestre_Capex = os.path.join(directorio_base, nombre_trimestre_actual, '2. Reacondicionamiento de Pozos (Workover CAPEX)')

crear_directorios(dfCo_trimestre, ruta_trimestre_CPI, 'CPI')
crear_directorios(dfC_trimestre, ruta_trimestre_Capex, 'CAPEX')
crear_directorios(dfO_trimestre, ruta_trimestre_Opex, 'OPEX')

# Crear el archivo Excel de Seguimiento del Reporte Trimestral
generar_informe_excel(estado_documentos, año_interes, trimestre_actual, directorio_base_Apoyo)

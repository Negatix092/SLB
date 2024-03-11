import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
import datetime

# se declara un diccionario de interés

estado_documentos = {'ACAK-189H': {'CPI': {'Propuesta Técnica': 'OK', 'Notificación': 'X', 'Reportes Diarios': 'X', 'Sumario': 'X', 'AIS': 'OK', None: {}, 'ARS': 'X'}}, 
                     'CNOF-056': {'Opex': {8: {'Propuesta Técnica': 'OK', 'Notificación': 'X', 'Reportes Diarios': 'X', 'Sumario': 'OK', 'AIS': 'OK', 'ARS': 'X'}}}, 
                     'CLBA-007R1': {'Opex': {6: {'Propuesta Técnica': 'OK', 'Notificación': 'OK', 'Reportes Diarios': 'OK, 12 archivos copiados', 'Sumario': 'X', 'AIS': 'OK', 'ARS': 'X'}}}, 
                     'CLBC-048H': {'Opex': {1: {'Propuesta Técnica': 'OK', 'Notificación': 'OK', 'Reportes Diarios': 'OK, 10 archivos copiados', 'Sumario': 'X', 'AIS': 'OK', 'ARS': 'X'}}}, 
                     'ACAE-066': {'Opex': {8: {'Propuesta Técnica': 'OK', 'Notificación': 'X', 'Reportes Diarios': 'X', 'Sumario': 'X', 'AIS': 'OK', 'ARS': 'X'}}}, 
                     'YLBD-029': {'Opex': {2: {'Propuesta Técnica': 'OK', 'Notificación': 'OK', 'Reportes Diarios': 'OK, 6 archivos copiados', 'Sumario': 'X', 'AIS': 'OK', 'Adecuación': 'Si', 'ARS': 'X'}}}, 
                     'ACAM-146': {'Opex': {3: {'Propuesta Técnica': 'OK', 'Notificación': 'OK', 'Reportes Diarios': 'OK, 12 archivos copiados', 'Sumario': 'X', 'AIS': 'OK', 'ARS': 'X'}}}, 
                     'CLBA-036I': {'Opex': {1: {'Propuesta Técnica': 'OK', 'Notificación': 'OK', 'Reportes Diarios': 'OK, 14 archivos copiados', 'Sumario': 'X', 'AIS': 'X', 'ARS': 'X'}}}, 
                     'CHSA-006': {'Opex': {4: {'Propuesta Técnica': 'OK', 'Notificación': 'OK', 'Reportes Diarios': 'OK, 7 archivos copiados', 'Sumario': 'X', 'AIS': 'OK', 'ARS': 'X'}}}, 
                     'CLBA-039H': {'Opex': {1: {'Propuesta Técnica': 'OK', 'Notificación': 'OK', 'Reportes Diarios': 'OK, 8 archivos copiados', 'Sumario': 'X', 'AIS': 'OK', 'ARS': 'X'}}}, 
                     'YCAE-028S1': {'Opex': {3: {'Propuesta Técnica': 'OK', 'Notificación': 'OK', 'Reportes Diarios': 'OK, 16 archivos copiados', 'Sumario': 'X', 'AIS': 'OK', 'ARS': 'X'}}}, 
                     'ACAF-086': {'Opex': {9: {'Propuesta Técnica': 'OK', 'Notificación': 'OK', 'Reportes Diarios': 'OK, 22 archivos copiados', 'Sumario': 'X', 'AIS': 'OK', 'ARS': 'X'}}}, 
                     'ACA-009': {'Opex': {17: {'Propuesta Técnica': 'X', 'Notificación': 'OK', 'Reportes Diarios': 'X', 'Sumario': 'OK', 'AIS': 'OK', 'ARS': 'X'}}}, 
                     'ACAF-163': {'Opex': {4: {'Propuesta Técnica': 'OK', 'Notificación': 'OK', 'Reportes Diarios': 'OK, 9 archivos copiados', 'Sumario': 'X', 'AIS': 'OK', 'ARS': 'X'}}}, 
                     'CNOC-013': {'Opex': {8: {'Propuesta Técnica': 'X', 'Notificación': 'X', 'Reportes Diarios': 'X', 'Sumario': 'X', 'AIS': 'OK', 'ARS': 'X'}}}, 
                     'ACAI-106R1': {'Capex': {1: {'Propuesta Técnica': 'OK', 'Notificación': 'OK', 'Reportes Diarios': 'OK, 21 archivos copiados', 'Sumario': 'X', 'AIS': 'OK', 'ARS': 'X'}}}, 
                     'YLBA-019': {'Capex': {1: {'Propuesta Técnica': 'X', 'Notificación': 'OK', 'Reportes Diarios': 'OK, 33 archivos copiados', 'Sumario': 'OK', 'AIS': 'OK', 'ARS': 'X'}}}}

def generar_informe_excel(estado_documentos):
    
    # se crea un dataframe a partir del diccionario
    data = []

    for pozo, pozo_info in estado_documentos.items():
        for tipo, tipo_info in pozo_info.items():
            adecuacion = None  # Valor por defecto para la adecuación
            if tipo == 'CPI':
                # Asegúrate de no incluir la clave None y su diccionario asociado {}
                cpi_info = {k: v for k, v in tipo_info.items() if k is not None}
                data.append([pozo, tipo, None] + list(cpi_info.values())) 
            else:
                for wo_number, values in tipo_info.items():
                    if 'Adecuación' in values:
                        adecuacion = values.pop('Adecuación')  # Remueve y guarda el valor de 'Adecuación'
                    # Asegúrate de incluir el valor de adecuación en la posición correcta
                    row = [pozo, tipo, wo_number] + list(values.values()) + [adecuacion]
                    data.append(row)

    # se crea el dataframe
    columns = ["Pozo", "Tipo", "WO #", "Propuesta Técnica", "Notificación", "Reportes Diarios", "Sumario", "AIS", "ARS", "Adecuación"]
    df = pd.DataFrame(data, columns=columns)

    # Obtener la fecha y hora actuales
    now = datetime.datetime.now()
    año_interes = now.year
    mes_actual = now.month
    trimestre_actual = (mes_actual - 1) // 3 + 1

    # Nombre del archivo basado en el trimestre y año actuales
    nombre_archivo = f'Informe Trimestral {trimestre_actual} {año_interes}.xlsx'
    ruta_guardado = fr'C:\Users\aperez143\OneDrive - SLB\Alex Perez\Reporte trimestral\Documentos\raw_code\{nombre_archivo}'

    # Guardar en un archivo excel
    with pd.ExcelWriter(ruta_guardado, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Ing. Prod-Eje.', index=False, startrow=1)
        
        # Obtenemos la hoja de trabajo de openpyxl para poder editarla
        workbook = writer.book
        worksheet = writer.sheets['Ing. Prod-Eje.']
        
        # Aplicar formato a las celdas
        for col in worksheet.columns:
            max_length = max((len(str(cell.value)) for cell in col)) + 2
            worksheet.column_dimensions[col[0].column_letter].width = max_length
        
        # Aplicar negrita a los títulos de las columnas y centrar todo el texto
        for cell in worksheet["2:2"]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Centrar todas las celdas
        for row in worksheet.iter_rows(min_row=1, max_col=len(df.columns), max_row=len(df)+2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
        
        # Agregar título sobre los nombres de las columnas
        worksheet.insert_rows(1)
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = f'Informe Trimestral {trimestre_actual} {año_interes}'
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        # Combinar las celdas para el título
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

# Ahora puedes llamar a la función generar_informe_excel pasando el diccionario de interés como argumento
generar_informe_excel(estado_documentos)
